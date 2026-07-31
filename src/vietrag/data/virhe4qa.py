"""Flexible ViRHE4QA parser.

The upstream archive has changed shape between releases. This parser reads
JSON, JSONL, and CSV sources and maps common field aliases into stable project
schemas. The original upstream train/dev/test field is intentionally ignored
for retrieval evaluation.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import zipfile
from pathlib import Path
from typing import Any, Iterable, Iterator

from openpyxl import load_workbook

from vietrag.data.preparation import prepare_document
from vietrag.preprocessing import safe_normalize
from vietrag.schemas import DocumentRecord, QueryRecord

_QUESTION_KEYS = ("question", "query", "cau_hoi", "Question")
_CONTEXT_KEYS = ("context", "passage", "paragraph", "Context")
_DOCUMENT_KEYS = (
    "document_name",
    "document",
    "law_name",
    "source_document",
    "Document",
)
_ARTICLE_KEYS = ("clause_name", "article_name", "article", "title", "Clause")
_ARTICLE_NUMBER_KEYS = ("article_number", "clause_number", "number")
_ANSWER_KEYS = (
    "abstractive_answer",
    "abstractive answer",
    "extractive_answer",
    "extractive answer",
    "answer",
    "answers",
    "Answer",
)


def _first(record: dict[str, Any], keys: Iterable[str], default: Any = "") -> Any:
    for key in keys:
        value = record.get(key)
        if value not in (None, "", []):
            return value
    return default


def _stable_token(prefix: str, value: str, length: int = 12) -> str:
    digest = hashlib.sha256(safe_normalize(value).encode("utf-8")).hexdigest()[:length]
    return f"{prefix}_{digest.upper()}"


def _answer_text(value: Any) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        if value and isinstance(value[0], dict):
            return str(value[0].get("text", value[0].get("answer", "")))
        return str(value[0]) if value else ""
    if isinstance(value, dict):
        text = value.get("text", value.get("answer", ""))
        return str(text[0] if isinstance(text, list) and text else text)
    return str(value or "")


def _iter_candidate_records(value: Any) -> Iterator[dict[str, Any]]:
    if isinstance(value, dict):
        if any(key in value for key in _QUESTION_KEYS) and any(
            key in value for key in _CONTEXT_KEYS
        ):
            yield value
        else:
            for nested in value.values():
                yield from _iter_candidate_records(nested)
    elif isinstance(value, list):
        for nested in value:
            yield from _iter_candidate_records(nested)


def _load_text_payload(name: str, payload: str) -> list[dict[str, Any]]:
    suffix = Path(name).suffix.lower()
    if suffix == ".json":
        return list(_iter_candidate_records(json.loads(payload)))
    if suffix == ".jsonl":
        return [
            record
            for line in payload.splitlines()
            if line.strip()
            for record in _iter_candidate_records(json.loads(line))
        ]
    if suffix == ".csv":
        return list(csv.DictReader(io.StringIO(payload)))
    return []


def _load_excel_payload(payload: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            values = worksheet.iter_rows(values_only=True)
            header = next(values, None)
            if header is None:
                continue
            keys = [str(value).strip() if value is not None else "" for value in header]
            for value_row in values:
                record = {
                    key: value
                    for key, value in zip(keys, value_row)
                    if key and value not in (None, "")
                }
                if record:
                    rows.append(record)
    finally:
        workbook.close()
    return rows


def load_source_rows(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path)
    if source.suffix.lower() == ".zip":
        rows: list[dict[str, Any]] = []
        with zipfile.ZipFile(source) as archive:
            for name in sorted(archive.namelist()):
                suffix = Path(name).suffix.lower()
                if suffix == ".xlsx":
                    workbook_rows = _load_excel_payload(archive.read(name))
                    upstream_split = Path(name).stem.lower()
                    for row in workbook_rows:
                        row["__upstream_split"] = upstream_split
                    rows.extend(workbook_rows)
                    continue
                if suffix not in {".json", ".jsonl", ".csv"}:
                    continue
                payload = archive.read(name).decode("utf-8-sig")
                rows.extend(_load_text_payload(name, payload))
        return rows
    if source.suffix.lower() == ".xlsx":
        return _load_excel_payload(source.read_bytes())
    with source.open("r", encoding="utf-8-sig", newline="") as stream:
        return _load_text_payload(source.name, stream.read())


def parse_rows(
    rows: Iterable[dict[str, Any]],
    *,
    source_name: str = "ViRHE4QA",
) -> tuple[list[DocumentRecord], list[QueryRecord]]:
    documents_by_article: dict[str, DocumentRecord] = {}
    queries: list[QueryRecord] = []
    for index, row in enumerate(rows, start=1):
        question = safe_normalize(str(_first(row, _QUESTION_KEYS)))
        context = str(_first(row, _CONTEXT_KEYS))
        if not question or not safe_normalize(context):
            continue
        document_name = safe_normalize(
            str(_first(row, _DOCUMENT_KEYS, "Unknown ViRHE4QA document"))
        )
        article_title = safe_normalize(
            str(_first(row, _ARTICLE_KEYS, "Untitled article"))
        )
        article_number = safe_normalize(
            str(_first(row, _ARTICLE_NUMBER_KEYS, ""))
        )
        doc_id = _stable_token("DOC", document_name)
        article_key = f"{document_name}|{article_title}|{context}"
        article_id = _stable_token(f"{doc_id}_ART", article_key)
        if article_id not in documents_by_article:
            upstream_split = str(
                row.get("__upstream_split", row.get("split", ""))
            )
            documents_by_article[article_id] = prepare_document(
                doc_id=doc_id,
                article_id=article_id,
                parent_id=doc_id,
                title=article_title,
                raw_text=context,
                source=source_name,
                metadata={
                    "document_name": document_name,
                    "article_number": article_number,
                    "source_url": None,
                    "upstream_split_ignored": upstream_split,
                    "upstream_splits_ignored": [upstream_split]
                    if upstream_split
                    else [],
                },
            )
        else:
            upstream_split = str(
                row.get("__upstream_split", row.get("split", ""))
            )
            provenance = documents_by_article[article_id].metadata.setdefault(
                "upstream_splits_ignored", []
            )
            if upstream_split and upstream_split not in provenance:
                provenance.append(upstream_split)
        query_id = _stable_token("Q", f"{index}|{question}|{article_id}")
        answer = _answer_text(_first(row, _ANSWER_KEYS, ""))
        queries.append(
            QueryRecord(
                query_id=query_id,
                base_query_id=query_id,
                query=question,
                clean_query=question,
                noise_type="clean",
                answerable=True,
                ambiguous=False,
                gold_context_ids=[article_id],
                reference_answer=safe_normalize(answer),
                source_dataset=source_name,
                split="unassigned",
            )
        )
    return sorted(documents_by_article.values(), key=lambda item: item.article_id), queries


def parse_virhe4qa(
    path: str | Path,
    *,
    source_name: str = "ViRHE4QA",
) -> tuple[list[DocumentRecord], list[QueryRecord], int]:
    rows = load_source_rows(path)
    documents, queries = parse_rows(rows, source_name=source_name)
    if not documents or not queries:
        raise ValueError(
            "No question/context records were found. Inspect the upstream schema "
            "with scripts/audit_data.py and extend the documented aliases if needed."
        )
    return documents, queries, len(rows)
